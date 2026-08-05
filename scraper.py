"""
Instagram Reel Scraper - core engine
=====================================
Scrapes data points (username, reel URL, music info, likes, comments, ...)
from Instagram Reel URLs using the USER'S OWN logged-in browser session
(cookies). The session is captured either by an interactive login flow or by
importing cookies exported from a browser extension.

Run the GUI:      python gui.py
Run the CLI:      python scraper.py --help
"""

from __future__ import annotations

import csv
import json
import queue
import random
import threading
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, List, Optional

from playwright.sync_api import (
    BrowserContext,
    Page,
    TimeoutError as PWTimeoutError,
    sync_playwright,
)

from parsers import (
    REEL_URL_RE,
    normalize_reel_url,
    parse_caption_from_html,
    parse_counts_from_html,
    parse_music_from_html,
    parse_uploaded_at_from_html,
    parse_username_from_html,
    parse_video_url_from_html,
)

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------

APP_DIR = Path(__file__).resolve().parent
DEFAULT_STATE_FILE = APP_DIR / "storage_state.json"

# Look like a real desktop Chrome to reduce bot-detection friction.
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)

BROWSER_ARGS = [
    "--disable-blink-features=AutomationControlled",
    "--disable-dev-shm-usage",
    "--no-first-run",
    "--no-default-browser-check",
]


def backoff_delay(consecutive_failures: int, base: float = 2.0, cap: float = 30.0) -> float:
    """Exponential cool-down (s) after N consecutive failed reels."""
    if consecutive_failures <= 0:
        return 0.0
    return min(cap, base * (2 ** (consecutive_failures - 1)))


def should_retry(status: str) -> bool:
    """Transient failures worth one automatic retry."""
    return status == "timeout" or status.startswith("error:")


# ----------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------


@dataclass
class ReelData:
    """One scraped reel record."""

    username: str = ""
    reel_url: str = ""
    music_title: str = ""
    music_artist: str = ""
    audio_page_url: str = ""
    caption: str = ""
    likes: str = ""
    comments: str = ""
    plays: str = ""
    uploaded_at: str = ""
    video_url: str = ""
    status: str = "ok"
    is_original_audio: bool = False

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def csv_columns() -> List[str]:
        return [
            "username",
            "reel_url",
            "music_title",
            "music_artist",
            "audio_page_url",
            "caption",
            "likes",
            "comments",
            "plays",
            "uploaded_at",
            "video_url",
            "status",
            "is_original_audio",
        ]


# ----------------------------------------------------------------------------
# Export helpers
# ----------------------------------------------------------------------------


def write_csv(results: List[ReelData], path: str | Path) -> None:
    cols = ReelData.csv_columns()
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        for r in results:
            writer.writerow(r.to_dict())


def export_json(results: List[ReelData], path: str | Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in results], f, ensure_ascii=False, indent=2)


def export_excel(results: List[ReelData], path: str | Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    cols = ReelData.csv_columns()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reels"
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0095F6")
    for r in results:
        d = r.to_dict()
        ws.append([d.get(c, "") for c in cols])

    widths = {
        "username": 22, "reel_url": 55, "music_title": 38, "music_artist": 24,
        "audio_page_url": 55, "caption": 60, "likes": 10, "comments": 12,
        "plays": 10, "uploaded_at": 22, "video_url": 60, "status": 14,
        "is_original_audio": 12,
    }
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 20)
    wb.save(path)


# ----------------------------------------------------------------------------
# Scraper engine
# ----------------------------------------------------------------------------


class InstagramReelScraper:
    def __init__(
        self,
        state_file: str | Path = DEFAULT_STATE_FILE,
        headless: bool = False,
        workers: int = 3,
        delay: float = 2.0,
        log: Optional[Callable[[str], None]] = None,
    ):
        self.state_file = Path(state_file)
        self.headless = headless
        self.workers = max(1, int(workers))
        self.delay = max(0.0, float(delay))
        self.log = log or (lambda msg: print(msg))
        self._stop = threading.Event()

    # ------------------------------------------------------------------ #
    # Session helpers
    # ------------------------------------------------------------------ #

    def has_session(self) -> bool:
        """True if a storage-state file with a (non-expired) sessionid exists."""
        if not self.state_file.exists():
            return False
        try:
            state = json.loads(self.state_file.read_text(encoding="utf-8"))
            cookies = state.get("cookies", [])
            now = time.time()
            return any(
                c.get("name") == "sessionid"
                and (
                    c.get("expires", -1) == -1
                    or float(c.get("expires", -1)) > now
                )
                for c in cookies
            )
        except Exception:
            return False

    def clear_session(self) -> None:
        if self.state_file.exists():
            self.state_file.unlink()

    def save_cookies_from_file(self, cookie_file: str | Path) -> bool:
        """
        Import cookies from a JSON file. Supports two formats:
          * Playwright storage_state JSON: {"cookies": [...], "origins": [...]}
          * EditThisCookie export:        [{"name":..., "value":..., ...}, ...]
        Returns True on success.
        """
        cookie_file = Path(cookie_file)
        raw = json.loads(cookie_file.read_text(encoding="utf-8"))

        if isinstance(raw, list):  # EditThisCookie style export
            same_site_map = {
                "lax": "Lax", "strict": "Strict", "none": "None",
                "no_restriction": "None", "unspecified": "Lax", "": "Lax",
            }
            cookies = []
            for c in raw:
                if not isinstance(c, dict) or "name" not in c:
                    continue
                domain = str(c.get("domain", "")).lstrip(".")
                if not domain:
                    continue
                same_site = c.get("sameSite")
                same_site_key = "" if same_site is None else str(same_site).lower()
                cookies.append(
                    {
                        "name": c["name"],
                        "value": str(c.get("value", "")),
                        "domain": "." + domain,
                        "path": c.get("path", "/"),
                        "expires": c.get("expirationDate", c.get("expires", -1)),
                        "httpOnly": c.get("httpOnly", False),
                        "secure": c.get("secure", False),
                        "sameSite": same_site_map.get(same_site_key, "Lax"),
                    }
                )
            self.state_file.write_text(
                json.dumps({"cookies": cookies, "origins": []}, indent=2),
                encoding="utf-8",
            )
            return True

        if isinstance(raw, dict) and "cookies" in raw:
            self.state_file.write_text(
                json.dumps(raw, indent=2), encoding="utf-8"
            )
            return True

        return False

    # ------------------------------------------------------------------ #
    # Login flow (captures the user's cookies into storage_state.json)
    # ------------------------------------------------------------------ #

    def login(self, timeout_seconds: int = 300) -> str:
        """
        Open a VISIBLE browser at instagram.com so the user can log in.
        As soon as Instagram issues a sessionid cookie we save the whole
        storage state locally and close the browser.
        Returns the logged-in username ('' if unknown / timed out).
        """
        with sync_playwright() as p:
            ctx_opts = {}
            if self.state_file.exists():
                try:
                    ctx_opts["storage_state"] = str(self.state_file)
                except Exception:
                    pass

            browser = p.chromium.launch(headless=False, args=BROWSER_ARGS)
            context = browser.new_context(
                **ctx_opts,
                user_agent=USER_AGENT,
                viewport={"width": 1280, "height": 900},
                locale="en-US",
            )
            page = context.new_page()
            page.goto(
                "https://www.instagram.com/accounts/login/",
                wait_until="domcontentloaded",
                timeout=60000,
            )
            self.log(
                "A browser window opened. Please log in to Instagram inside it. "
                "The session is saved automatically once you are logged in."
            )

            username = ""
            deadline = time.time() + timeout_seconds
            while time.time() < deadline:
                if self._stop.is_set():
                    break
                cookies = {
                    c["name"]: c["value"]
                    for c in context.cookies("https://www.instagram.com")
                }
                if cookies.get("sessionid"):
                    context.storage_state(path=str(self.state_file))
                    try:
                        page.goto(
                            "https://www.instagram.com/",
                            wait_until="domcontentloaded",
                            timeout=30000,
                        )
                        page.wait_for_timeout(1500)
                        username = page.evaluate(
                            """
                            () => {
                                const img = document.querySelector(
                                    'a[href^="/"] img[alt], nav img[alt], main img[alt]'
                                );
                                if (!img) return '';
                                return (img.getAttribute('alt') || '')
                                    .replace(/\u2019s profile picture/gi, '')
                                    .replace(/'s profile picture/gi, '')
                                    .trim();
                            }
                            """
                        ) or ""
                    except Exception:
                        pass
                    self.log(
                        f"[OK] Session saved to {self.state_file.name} "
                        f"(user: {username or 'unknown'})"
                    )
                    browser.close()
                    return username
                page.wait_for_timeout(2000)

            browser.close()
            self.log("[!] Login timed out - no session was saved.")
            return ""

    # ------------------------------------------------------------------ #
    # Scraping
    # ------------------------------------------------------------------ #

    def scrape(
        self,
        urls: List[str],
        progress_cb: Optional[Callable[[int, int], None]] = None,
    ) -> List[ReelData]:
        """
        Scrape every URL. `workers` parallel browser windows are used, one
        URL at a time each, re-using the saved user session (cookies).
        """
        self._stop.clear()
        results: List[ReelData] = []

        if not self.has_session():
            self.log(
                "[!] No valid saved session found. Click 'Login to Instagram' "
                "first (or import a cookies file)."
            )

        q: "queue.Queue[str]" = queue.Queue()
        seen = set()
        for u in urls:
            u = u.strip()
            if u and u not in seen:
                seen.add(u)
                q.put(u)

        total = q.qsize()
        if total == 0:
            return results

        n_workers = min(self.workers, total)
        self.log(
            f"[>] Starting scrape of {total} reel(s) with "
            f"{n_workers} parallel worker(s)..."
        )
        if not self.headless:
            self.log(
                "[>] Browser windows will open visibly. "
                "You can watch the scrape happen."
            )

        lock = threading.Lock()

        def worker_fn(wid: int) -> None:
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(
                        headless=self.headless, args=BROWSER_ARGS
                    )
                    context = browser.new_context(
                        storage_state=str(self.state_file)
                        if self.state_file.exists()
                        else None,
                        user_agent=USER_AGENT,
                        viewport={"width": 1280, "height": 900},
                        locale="en-US",
                    )
                    fail_streak = 0
                    while True:
                        try:
                            url = q.get_nowait()
                        except queue.Empty:
                            break
                        if self._stop.is_set():
                            break
                        self.log(f"    worker {wid} -> {url}")
                        page = context.new_page()
                        try:
                            data = self._scrape_one(page, url)
                        finally:
                            page.close()
                        with lock:
                            results.append(data)
                        if progress_cb:
                            progress_cb(len(results), total)
                        fail_streak = fail_streak + 1 if data.status != "ok" else 0
                        wait = (
                            self.delay
                            + random.uniform(0, 1.5)
                            + backoff_delay(fail_streak)
                        )
                        if fail_streak >= 2:
                            self.log(
                                f"    worker {wid}: {fail_streak} failures in a row, "
                                f"cooling down {wait:.1f}s"
                            )
                        time.sleep(wait)
                    context.close()
                    browser.close()
            except Exception as e:  # pragma: no cover - defensive
                self.log(f"    [x] worker {wid} crashed: {e}")

        threads = [
            threading.Thread(target=worker_fn, args=(i + 1,), daemon=True)
            for i in range(n_workers)
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self.log(f"[OK] Done. {len(results)}/{total} scraped.")
        return results

    def stop(self) -> None:
        self._stop.set()

    # ------------------------------------------------------------------ #
    # Single page scrape
    # ------------------------------------------------------------------ #

    def _scrape_one(self, page: Page, url: str) -> ReelData:
        data = ReelData(reel_url=url)
        for attempt in (1, 2):
            data = self._attempt_one(page, url)
            if data.status == "ok" or not should_retry(data.status):
                break
            if self._stop.is_set():
                break
            self.log(f"    retry {url} (attempt {attempt + 1})")
            page.wait_for_timeout(1500)
        return data

    def _attempt_one(self, page: Page, url: str) -> ReelData:
        """Single navigation + extraction pass for one reel URL."""
        data = ReelData(reel_url=url)
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            self._dismiss_overlays(page)

            if self._is_login_wall(page):
                data.status = "session_expired"
                self.log(
                    f"    [x] {url} - session expired / login wall. "
                    "Re-login via the Login button."
                )
                return data

            try:
                page.wait_for_selector(
                    'a[href*="/reels/audio/"], meta[property="og:title"], '
                    "article, video",
                    timeout=30000,
                )
            except PWTimeoutError:
                pass

            page.wait_for_timeout(2500)  # let lazy content settle
            data = self._extract(page)
            if not data.reel_url:
                data.reel_url = url  # keep the input URL if canonical missing

            if self._is_unavailable(page):
                data.status = "unavailable"
                return data

            # One reload retry if the page rendered without core data.
            if not data.username and not data.music_title:
                try:
                    page.reload(wait_until="domcontentloaded", timeout=45000)
                    page.wait_for_timeout(2500)
                    data = self._extract(page)
                except Exception:
                    pass

            data.status = "ok"
        except PWTimeoutError:
            data.status = "timeout"
            self.log(f"    [x] {url} - timed out.")
        except Exception as e:  # pragma: no cover - defensive
            data.status = f"error: {type(e).__name__}"
            self.log(f"    [x] {url} -> {e}")
        return data

    # ------------------------------------------------------------------ #
    # Extraction
    # ------------------------------------------------------------------ #

    def _extract(self, page: Page) -> ReelData:
        d = ReelData()
        try:
            html = page.content()
        except Exception:
            html = ""

        # canonical reel URL
        try:
            d.reel_url = page.eval_on_selector(
                'link[rel="canonical"]', "el => el.href"
            ) or d.reel_url
        except Exception:
            pass

        # --- username: parsers first, article-header anchor as fallback ---
        d.username = parse_username_from_html(html)
        if not d.username:
            try:
                d.username = page.evaluate(
                    """
                    () => {
                        const anchors = document.querySelectorAll(
                            'article header a[href^="/"], article a[href^="/"]'
                        );
                        const re = /^\/([\w.]+)\/?$/;
                        for (const a of anchors) {
                            const m = (a.getAttribute('href') || '').match(re);
                            if (m) return m[1];
                        }
                        return '';
                    }
                    """
                ) or ""
            except Exception:
                pass

        # --- music info ---
        music = parse_music_from_html(html)
        d.music_title = music["title"]
        d.music_artist = music["artist"]
        d.audio_page_url = music["audio_page_url"]
        d.is_original_audio = music["original"]

        # --- likes / comments / plays: parsers first, DOM spans as fallback ---
        counts = parse_counts_from_html(html)
        d.likes, d.comments, d.plays = counts["likes"], counts["comments"], counts["plays"]
        if not d.likes or not d.comments:
            try:
                dom_counts = page.evaluate(
                    r"""
                    () => {
                        const pick = (sel) => {
                            const el = document.querySelector(sel);
                            return el ? (el.textContent || '')
                                .replace(/\s+/g, ' ').trim() : '';
                        };
                        return {
                            likes: pick('a[href*="/liked_by/"] span'),
                            comments: pick('a[href*="/comments/"] span'),
                        };
                    }
                    """
                ) or {}
                d.likes = d.likes or (dom_counts.get("likes") or "")
                d.comments = d.comments or (dom_counts.get("comments") or "")
            except Exception:
                pass

        # --- caption / upload date / video URL ---
        d.caption = parse_caption_from_html(html)
        d.uploaded_at = parse_uploaded_at_from_html(html)
        d.video_url = parse_video_url_from_html(html)
        return d

    # ------------------------------------------------------------------ #
    # Page-state checks / helpers
    # ------------------------------------------------------------------ #

    def _is_login_wall(self, page: Page) -> bool:
        try:
            if "/accounts/login" in page.url:
                return True
            return bool(
                page.evaluate(
                    """
                    () => !!document.querySelector(
                        'input[name="username"], input[name="password"]'
                    )
                    """
                )
            )
        except Exception:
            return False

    def _is_unavailable(self, page: Page) -> bool:
        try:
            body = page.evaluate(
                "() => document.body ? document.body.innerText.slice(0, 1500) : ''"
            )
            return (
                "isn't available" in body
                or "This page isn't available" in body
                or "Sorry, this page isn't available" in body
            )
        except Exception:
            return False

    def _dismiss_overlays(self, page: Page) -> None:
        """Best-effort click-through of cookie / login-suggest dialogs."""
        selectors = [
            'div[role="dialog"] button:has-text("Allow")',
            'div[role="dialog"] button:has-text("Accept")',
            'button:has-text("Not now")',
            'button[aria-label="Close"]',
        ]
        for sel in selectors:
            try:
                btn = page.locator(sel).first
                if btn.count() and btn.is_visible():
                    btn.click(timeout=1500)
            except Exception:
                pass


# ----------------------------------------------------------------------------
# CLI entry point
# ----------------------------------------------------------------------------


def _print_table(results: List[ReelData]) -> None:
    if not results:
        return
    print(f"\n{'USERNAME':<20} {'MUSIC':<40} {'STATUS':<18} REEL URL")
    print("-" * 110)
    for r in results:
        print(
            f"{r.username[:19]:<20} {(r.music_title or '')[:39]:<40} "
            f"{r.status:<18} {r.reel_url}"
        )


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Scrape Instagram Reels using your own logged-in session."
    )
    parser.add_argument("urls", nargs="*", help="one or more reel URLs")
    parser.add_argument(
        "-f", "--file", help="text file with one reel URL per line"
    )
    parser.add_argument(
        "--login", action="store_true",
        help="open a browser so you can log in and save the session",
    )
    parser.add_argument(
        "--import-cookies", metavar="FILE",
        help="import cookies from a JSON file (Playwright state or "
        "EditThisCookie export) and exit",
    )
    parser.add_argument(
        "--clear-session", action="store_true",
        help="delete the saved session file and exit",
    )
    parser.add_argument(
        "--headless", action="store_true",
        help="run browsers invisibly (may trigger IG login walls)",
    )
    parser.add_argument(
        "-w", "--workers", type=int, default=3,
        help="number of parallel browser workers (default: 3)",
    )
    parser.add_argument(
        "--delay", type=float, default=2.0,
        help="pause between reels in seconds (default: 2.0)",
    )
    parser.add_argument(
        "--state", default=str(DEFAULT_STATE_FILE),
        help="path to the session storage-state file",
    )
    parser.add_argument(
        "-o", "--output", default="reels_results.csv",
        help="output CSV path (default: reels_results.csv)",
    )
    args = parser.parse_args()

    scraper = InstagramReelScraper(
        state_file=args.state,
        headless=args.headless,
        workers=args.workers,
        delay=args.delay,
    )

    if args.clear_session:
        scraper.clear_session()
        print("[OK] Session cleared.")
        return

    if args.import_cookies:
        if scraper.save_cookies_from_file(args.import_cookies):
            print(f"[OK] Cookies imported -> {scraper.state_file}")
        else:
            print("[x] Could not read that cookie file.")
        return

    if args.login:
        scraper.login()
        return

    urls = list(args.urls)
    if args.file:
        urls += [
            line.strip()
            for line in Path(args.file).read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]

    normalized = []
    for u in urls:
        norm = normalize_reel_url(u)
        if norm:
            normalized.append(norm)
        else:
            print(f"[!] Skipped (not a reel URL): {u}")
    urls = normalized

    if not urls:
        parser.print_help()
        return

    results = scraper.scrape(urls)
    if results:
        write_csv(results, args.output)
        print(f"[OK] Results written to {args.output}")
    _print_table(results)


if __name__ == "__main__":
    main()
